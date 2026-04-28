#!/usr/bin/env python3
"""
Read Essential Bali article-tracker xlsx.

Inbox: /opt/.openclaw-ess/inbox/articles/*.xlsx (synced from local).

Sheets are month tabs (Apr, May, June, July, ...) with these columns:
  No · Client · Site · Blog Topic · SEO Title · Meta Description · Keywords ·
  Draft Link · Writer · Progress · Article Status · QA · Published Link ·
  Content Type · Notes

Outputs JSON list of usable rows (Article Status != Published, has SEO Title).
Each row maps to a Payload Articles draft.

Usage:
  ./read-articles-xlsx.py [path]            # default: latest in inbox
  ./read-articles-xlsx.py --status=draft    # filter
  ./read-articles-xlsx.py --month=Apr       # one sheet
"""
from __future__ import annotations
import argparse
import json
import os
import re
import sys
from glob import glob
from pathlib import Path

INBOX = Path("/opt/.openclaw-ess/inbox/articles")

COL = {
    "no": 0,
    "client": 1,
    "site": 2,
    "blog_topic": 3,
    "seo_title": 4,
    "meta_description": 5,
    "keywords": 6,
    "draft_link": 7,
    "writer": 8,
    "progress": 9,
    "article_status": 10,
    "qa": 11,
    "published_link": 12,
    "content_type": 13,
    "notes": 14,
}

# Heuristic mapping topic phrases → Essential Bali topic slugs.
TOPIC_HINTS = [
    ("event", "events"),
    ("news", "news"),
    ("dine", "dine"),
    ("restaurant", "dine"),
    ("eat", "dine"),
    ("food", "dine"),
    ("warung", "dine"),
    ("yoga", "health-wellness"),
    ("wellness", "health-wellness"),
    ("retreat", "health-wellness"),
    ("spa", "health-wellness"),
    ("nightlife", "nightlife"),
    ("bar", "nightlife"),
    ("club", "nightlife"),
    ("dj", "nightlife"),
    ("activit", "activities"),
    ("surf", "activities"),
    ("dive", "activities"),
    ("trek", "activities"),
    ("snorkel", "activities"),
    ("temple", "people-culture"),
    ("ceremony", "people-culture"),
    ("artisan", "people-culture"),
    ("culture", "people-culture"),
    ("community", "people-culture"),
    ("featured", "featured"),
]

AREAS = [
    "canggu",
    "kuta",
    "ubud",
    "jimbaran",
    "denpasar",
    "kintamani",
    "singaraja",
    "nusa penida",
]


def slugify(s: str) -> str:
    s = re.sub(r"[^\w\s-]", "", s.lower()).strip()
    s = re.sub(r"[\s_]+", "-", s)
    return s[:80]


def guess_area(text: str) -> str | None:
    t = text.lower()
    for a in AREAS:
        if a in t:
            return a.replace(" ", "-")
    return None


def guess_topic(text: str) -> str | None:
    t = text.lower()
    for phrase, slug in TOPIC_HINTS:
        if phrase in t:
            return slug
    return None


def latest_xlsx() -> Path | None:
    if not INBOX.exists():
        return None
    files = sorted(INBOX.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def read(path: Path, status_filter: str | None, month_filter: str | None):
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed; run: pip install openpyxl"}), file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        if sheet.lower() == "dashboard":
            continue
        if month_filter and sheet.lower() != month_filter.lower():
            continue
        ws = wb[sheet]
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r == 0:
                continue
            if not row or len(row) < 5:
                continue
            seo_title = (row[COL["seo_title"]] or "").strip() if row[COL["seo_title"]] else ""
            if not seo_title:
                continue
            article_status = (row[COL["article_status"]] or "").strip().lower() if row[COL["article_status"]] else ""
            if status_filter and status_filter.lower() not in article_status:
                continue
            blog_topic = row[COL["blog_topic"]] or ""
            seed_text = f"{seo_title} {blog_topic}"
            keywords_raw = row[COL["keywords"]] or ""
            keywords = [k.strip() for k in re.split(r"[,;]", str(keywords_raw)) if k.strip()]
            out.append(
                {
                    "month": sheet,
                    "row": r + 1,
                    "no": row[COL["no"]],
                    "title": seo_title,
                    "slug": slugify(seo_title),
                    "topic_brief": str(blog_topic).strip(),
                    "meta_description": (row[COL["meta_description"]] or "").strip() if row[COL["meta_description"]] else "",
                    "keywords": keywords,
                    "draft_link": (row[COL["draft_link"]] or "").strip() if row[COL["draft_link"]] else "",
                    "writer": (row[COL["writer"]] or "").strip() if row[COL["writer"]] else "",
                    "article_status": article_status,
                    "guessed_area": guess_area(seed_text),
                    "guessed_topic": guess_topic(seed_text),
                }
            )
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path", nargs="?")
    ap.add_argument("--status", help="filter by Article Status substring (e.g. draft, published)")
    ap.add_argument("--month", help="single month sheet (Apr, May, ...)")
    ap.add_argument("--out", help="write JSON to file instead of stdout")
    args = ap.parse_args()

    path = Path(args.path) if args.path else latest_xlsx()
    if not path or not path.exists():
        print(json.dumps({"error": f"no xlsx at {path or INBOX}"}), file=sys.stderr)
        sys.exit(1)

    rows = read(path, args.status, args.month)
    payload = {"file": str(path), "count": len(rows), "rows": rows}
    text = json.dumps(payload, indent=2, ensure_ascii=False)
    if args.out:
        Path(args.out).write_text(text)
        print(f"wrote {args.out}: {len(rows)} rows", file=sys.stderr)
    else:
        print(text)


if __name__ == "__main__":
    main()
